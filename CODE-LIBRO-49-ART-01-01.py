#!/usr/bin/env python3
"""Integrated Climate Futures and Decision Support for Resilient Sugarcane.

This program builds and executes the synthetic worked case used in the chapter of
that title. It couples a balanced climate-scenario ensemble with a transparent
monthly crop-water model, economic and sustainability accounting, machine-learning
emulation, explainability, distribution-free predictive intervals, global
sensitivity analysis, robust multicriteria decision analysis, and adaptive pathways.

The outputs are organized in three folders that correspond exactly to Sections 5.1,
5.2, and 5.3 of the chapter. The numerical values are synthetic stress-test results;
they are not prescriptions for a real production region and require local calibration
before operational use.
"""
from __future__ import annotations

import argparse
import hashlib
import importlib
import json
import math
import os
import subprocess
import sys
import textwrap
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Tuple

# -----------------------------------------------------------------------------
# Dependency verification
# -----------------------------------------------------------------------------

REQUIRED_PACKAGES: Mapping[str, str] = {
    "numpy": "numpy",
    "pandas": "pandas",
    "scipy": "scipy",
    "sklearn": "scikit-learn",
    "matplotlib": "matplotlib",
    "openpyxl": "openpyxl",
    "xgboost": "xgboost",
    "shap": "shap",
    "networkx": "networkx",
}


def verify_dependencies(auto_install: bool = False) -> None:
    """Check required libraries and optionally install missing packages."""
    missing: List[str] = []
    for module_name, package_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing.append(package_name)
    if missing and auto_install:
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
        missing = []
        for module_name, package_name in REQUIRED_PACKAGES.items():
            try:
                importlib.import_module(module_name)
            except ImportError:
                missing.append(package_name)
    if missing:
        install_command = f'{sys.executable} -m pip install ' + " ".join(missing)
        raise RuntimeError(
            "Missing required packages: " + ", ".join(missing) +
            "\nInstall them with:\n" + install_command +
            "\nAlternatively, rerun this program with --auto-install."
        )


# Verify before importing the scientific stack.
verify_dependencies("--auto-install" in sys.argv)

import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import pandas as pd
import shap
from matplotlib.patches import FancyBboxPatch
from matplotlib.ticker import PercentFormatter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from scipy import stats
from scipy.stats import qmc
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import ExtraTreesRegressor, RandomForestRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from xgboost import XGBRegressor


# -----------------------------------------------------------------------------
# Configuration and model specification
# -----------------------------------------------------------------------------

@dataclass(frozen=True)
class ResearchConfig:
    random_seed: int = 4901
    n_scenarios: int = 1800
    n_weight_samples: int = 3000
    sobol_base_size: int = 512
    planning_horizon_years: int = 15
    baseline_potential_yield_t_ha: float = 108.0
    soil_water_capacity_mm: float = 185.0
    initial_soil_water_fraction: float = 0.70
    target_conformal_coverage: float = 0.90
    figure_dpi: int = 300
    output_dir: str = ""


@dataclass(frozen=True)
class Strategy:
    strategy_id: int
    name: str
    short_name: str
    potential_yield_factor: float
    heat_tolerance_shift_c: float
    water_use_efficiency_factor: float
    irrigation_capacity_mm_year: float
    irrigation_efficiency: float
    evaporation_reduction: float
    planting_shift_months: int
    storage_capture_fraction: float
    drainage_protection: float
    pest_protection: float
    capex_usd_ha: float
    annual_opex_usd_ha: float
    energy_bonus_mwh_ha: float
    ghg_reduction_tco2e_ha: float
    labor_change_days_ha: float
    equity_bonus: float
    ecological_bonus: float
    implementation_complexity: float
    minimum_institutional_capacity: float


STRATEGIES: Tuple[Strategy, ...] = (
    Strategy(0, "Current practice", "Current", 1.000, 0.0, 1.00, 90, 0.62, 0.00, 0, 0.00, 0.05, 0.00, 0, 0, 0.00, 0.00, 0.0, 0.00, 0.00, 0.20, 0.20),
    Strategy(1, "Drought-tolerant cultivar", "Cultivar", 1.035, 0.7, 1.10, 90, 0.62, 0.00, 0, 0.00, 0.05, 0.08, 520, 65, 0.00, 0.10, 1.0, 0.06, 0.04, 0.40, 0.35),
    Strategy(2, "Regulated deficit irrigation", "Deficit irrigation", 1.010, 0.2, 1.18, 240, 0.86, 0.00, 0, 0.00, 0.05, 0.00, 1450, 175, 0.00, -0.05, -1.0, 0.00, 0.08, 0.58, 0.45),
    Strategy(3, "Soil-moisture conservation", "Soil conservation", 1.025, 0.2, 1.08, 100, 0.64, 0.19, 0, 0.08, 0.08, 0.03, 690, 95, 0.00, 0.28, 2.0, 0.10, 0.18, 0.38, 0.30),
    Strategy(4, "Climate-adjusted planting window", "Planting shift", 1.015, 0.4, 1.03, 100, 0.64, 0.03, -1, 0.00, 0.08, 0.02, 330, 35, 0.00, 0.04, 0.5, 0.05, 0.05, 0.32, 0.28),
    Strategy(5, "On-farm water storage", "Water storage", 1.015, 0.2, 1.08, 330, 0.80, 0.02, 0, 0.32, 0.10, 0.00, 3220, 210, 0.00, -0.08, -0.5, -0.03, 0.05, 0.76, 0.62),
    Strategy(6, "Integrated cultivar-irrigation-soil package", "Integrated package", 1.075, 1.0, 1.27, 250, 0.90, 0.22, -1, 0.16, 0.12, 0.10, 2940, 315, 0.18, 0.48, 2.5, 0.10, 0.24, 0.72, 0.58),
    Strategy(7, "Drainage and harvest coordination", "Harvest coordination", 1.025, 0.2, 1.02, 95, 0.64, 0.00, 0, 0.00, 0.70, 0.03, 1180, 120, 0.10, 0.20, -1.5, 0.08, 0.10, 0.56, 0.48),
    Strategy(8, "Circular climate-smart production package", "Circular package", 1.085, 1.1, 1.28, 260, 0.91, 0.24, -1, 0.20, 0.25, 0.12, 4180, 390, 1.05, 1.18, 3.0, 0.15, 0.32, 0.84, 0.68),
)

PATHWAY_HORIZON_PARAMETERS: Mapping[Tuple[str, int], Tuple[float, float, float, float]] = {
    ("SSP1-2.6", 2035): (1.05, 0.00, 11.0, 445.0),
    ("SSP1-2.6", 2055): (1.45, -0.025, 15.5, 475.0),
    ("SSP1-2.6", 2075): (1.55, -0.030, 17.5, 495.0),
    ("SSP2-4.5", 2035): (1.35, -0.025, 14.5, 455.0),
    ("SSP2-4.5", 2055): (2.15, -0.060, 24.5, 515.0),
    ("SSP2-4.5", 2075): (2.85, -0.095, 33.0, 575.0),
    ("SSP5-8.5", 2035): (1.60, -0.040, 18.0, 470.0),
    ("SSP5-8.5", 2055): (2.80, -0.100, 33.5, 600.0),
    ("SSP5-8.5", 2075): (4.15, -0.160, 51.0, 720.0),
}

BASE_MONTHLY_TEMPERATURE_C = np.array([23.6, 24.5, 25.7, 27.0, 28.1, 28.6, 28.2, 27.7, 27.1, 26.2, 24.9, 23.8])
BASE_MONTHLY_PRECIPITATION_MM = np.array([34, 28, 31, 49, 92, 157, 186, 174, 155, 105, 55, 38], dtype=float)
BASE_MONTHLY_ET0_MM = np.array([100, 106, 121, 132, 139, 141, 137, 132, 123, 113, 104, 98], dtype=float)
CROP_COEFFICIENT = np.array([0.48, 0.58, 0.78, 0.96, 1.10, 1.18, 1.23, 1.21, 1.12, 0.98, 0.81, 0.62])
STAGE_YIELD_WEIGHTS = np.array([0.025, 0.035, 0.065, 0.100, 0.130, 0.145, 0.150, 0.140, 0.105, 0.060, 0.030, 0.015])


def default_output_dir() -> Path:
    desktop = Path.home() / "Desktop"
    base = desktop if desktop.exists() else Path.cwd()
    return base / "Integrated_Climate_Futures_Sugarcane_Results"


def ensure_output_structure(root: Path) -> Dict[str, Path]:
    paths = {
        "root": root,
        "data": root / "data",
        "s51": root / "5.1_Climate_and_System_Response",
        "s52": root / "5.2_Predictive_Performance_and_Uncertainty",
        "s53": root / "5.3_Robust_Decisions_and_Adaptive_Pathways",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def minmax(values: pd.Series | np.ndarray, benefit: bool = True) -> pd.Series:
    series = pd.Series(values, copy=False, dtype=float)
    lo, hi = float(series.min()), float(series.max())
    normalized = pd.Series(0.5, index=series.index) if math.isclose(lo, hi) else (series - lo) / (hi - lo)
    return normalized if benefit else 1.0 - normalized


def save_figure(fig: plt.Figure, path: Path, dpi: int) -> None:
    fig.tight_layout()
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)


# -----------------------------------------------------------------------------
# Scenario ensemble
# -----------------------------------------------------------------------------

def normalized_severity(temp: np.ndarray, precip: np.ndarray, heat: np.ndarray, rain_cv: np.ndarray) -> np.ndarray:
    return np.clip(
        0.42 * np.clip((temp - 0.6) / 4.8, 0, 1)
        + 0.28 * np.clip((-precip + 0.01) / 0.34, 0, 1)
        + 0.20 * np.clip((heat - 5.0) / 52.0, 0, 1)
        + 0.10 * np.clip((rain_cv - 0.15) / 0.31, 0, 1),
        0,
        1,
    )


def generate_scenario_ensemble(config: ResearchConfig) -> pd.DataFrame:
    cells = list(PATHWAY_HORIZON_PARAMETERS)
    if config.n_scenarios % len(cells) != 0:
        raise ValueError("n_scenarios must be divisible by nine to balance pathway-horizon cells.")
    per_cell = config.n_scenarios // len(cells)
    records: List[pd.DataFrame] = []
    scenario_start = 0
    for cell_index, (pathway, horizon) in enumerate(cells):
        temp_mean, precip_mean, heat_mean, co2_mean = PATHWAY_HORIZON_PARAMETERS[(pathway, horizon)]
        sampler = qmc.LatinHypercube(d=12, seed=config.random_seed + 29 * cell_index)
        u = sampler.random(per_cell)
        z = stats.norm.ppf(np.clip(u, 1e-7, 1 - 1e-7))
        temp_sd = 0.20 + 0.055 * temp_mean
        precip_sd = 0.045 + 0.018 * temp_mean
        heat_sd = 3.2 + 0.70 * temp_mean
        temperature = np.clip(temp_mean + temp_sd * z[:, 0], 0.55, 5.6)
        precipitation = np.clip(precip_mean + precip_sd * z[:, 1], -0.36, 0.16)
        heatwaves = np.clip(heat_mean + heat_sd * z[:, 2], 4, 72)
        co2 = np.clip(co2_mean + (12 + 4 * temp_mean) * z[:, 3], 410, 830)
        rain_cv = np.clip(0.22 + 0.035 * temp_mean + 0.055 * z[:, 4], 0.14, 0.49)
        institutional = np.clip(0.58 + 0.17 * z[:, 5], 0.22, 0.96)
        adoption = np.clip(0.28 + 0.68 * institutional + 0.055 * z[:, 6], 0.25, 0.98)
        sugar_price = np.clip(1.00 + 0.16 * z[:, 7] + 0.035 * temp_mean, 0.72, 1.48)
        energy_price = np.clip(1.00 + 0.17 * z[:, 8] + 0.025 * temp_mean, 0.72, 1.55)
        input_cost = np.clip(1.00 + 0.13 * z[:, 9] + 0.045 * temp_mean, 0.80, 1.55)
        water_cost = np.clip(0.058 + 0.013 * temp_mean + 0.020 * z[:, 10], 0.025, 0.145)
        pest = np.clip(0.07 + 0.034 * temp_mean + 0.055 * z[:, 11], 0.01, 0.40)
        flood = np.clip(0.35 * rain_cv + 0.50 * np.maximum(precipitation, 0) + 0.04, 0.02, 0.55)
        discount = np.clip(0.104 - 0.050 * institutional + 0.007 * z[:, 9], 0.04, 0.11)
        structure = np.clip(1.0 + 0.027 * z[:, 6], 0.92, 1.08)
        severity = normalized_severity(temperature, precipitation, heatwaves, rain_cv)
        records.append(pd.DataFrame({
            "scenario_id": np.arange(scenario_start, scenario_start + per_cell),
            "pathway": pathway,
            "horizon": horizon,
            "temperature_anomaly_c": temperature,
            "precipitation_change_fraction": precipitation,
            "heatwave_days": heatwaves,
            "co2_ppm": co2,
            "rainfall_cv": rain_cv,
            "sugar_price_index": sugar_price,
            "energy_price_index": energy_price,
            "input_cost_index": input_cost,
            "water_cost_usd_m3": water_cost,
            "institutional_capacity": institutional,
            "adoption_fraction": adoption,
            "discount_rate": discount,
            "pest_pressure": pest,
            "flood_index": flood,
            "model_structure_factor": structure,
            "climate_severity_index": severity,
        }))
        scenario_start += per_cell
    return pd.concat(records, ignore_index=True).sort_values("scenario_id").reset_index(drop=True)


# -----------------------------------------------------------------------------
# Vectorized monthly crop-water-economic model
# -----------------------------------------------------------------------------

def strategy_frame() -> pd.DataFrame:
    return pd.DataFrame([asdict(s) for s in STRATEGIES])


def expand_scenarios_and_strategies(scenarios: pd.DataFrame) -> pd.DataFrame:
    a = scenarios.assign(_join_key=1)
    b = strategy_frame().assign(_join_key=1)
    return a.merge(b, on="_join_key", how="inner").drop(columns="_join_key")


def simulate_expanded(expanded: pd.DataFrame, config: ResearchConfig) -> pd.DataFrame:
    n = len(expanded)
    # Convert all frequently used columns once.
    def col(name: str) -> np.ndarray:
        return expanded[name].to_numpy(dtype=float)

    temp_anom = col("temperature_anomaly_c")
    precip_change = col("precipitation_change_fraction")
    rain_cv = col("rainfall_cv")
    institutional = col("institutional_capacity")
    adoption = col("adoption_fraction")
    complexity = col("implementation_complexity")
    minimum_capacity = col("minimum_institutional_capacity")
    feasibility_gap = np.maximum(0.0, minimum_capacity - institutional)
    implementation = np.clip(adoption * (1.0 - 0.85 * feasibility_gap), 0.16, 1.0)

    soil_capacity = config.soil_water_capacity_mm
    soil_water = np.full(n, soil_capacity * config.initial_soil_water_fraction)
    carried_storage = np.zeros(n)
    potential_et = np.zeros(n)
    actual_et = np.zeros(n)
    irrigation_applied = np.zeros(n)
    effective_precipitation = np.zeros(n)
    runoff_total = np.zeros(n)
    monthly_satisfaction = np.zeros((n, 12))

    irrigation_capacity = col("irrigation_capacity_mm_year") * (0.62 + 0.38 * institutional)
    irrigation_budget = irrigation_capacity / 12.0
    irrigation_efficiency = col("irrigation_efficiency")
    storage_capture = col("storage_capture_fraction")
    drainage = col("drainage_protection")
    ecological_bonus = col("ecological_bonus")
    evap_reduction = col("evaporation_reduction")
    planting_shift = expanded["planting_shift_months"].to_numpy(dtype=int)

    # Monthly process loop is vectorized across all scenario-strategy combinations.
    for month in range(12):
        shifted_month = (month - planting_shift) % 12
        kc = CROP_COEFFICIENT[shifted_month]
        seasonal_dryness = 1.0 + rain_cv * np.sin((month + 1) * 2 * np.pi / 12.0 + 0.7)
        raw_rain = BASE_MONTHLY_PRECIPITATION_MM[month] * (1.0 + precip_change) * seasonal_dryness
        raw_rain = np.maximum(raw_rain, 0.0)
        flood_loss = np.maximum(raw_rain - 165.0, 0.0) * (0.55 - 0.40 * drainage)
        effective_rain = np.maximum(0.0, raw_rain * 0.84 - flood_loss)
        runoff = np.maximum(0.0, raw_rain - effective_rain) * (0.62 - 0.28 * ecological_bonus)
        wet = raw_rain > 125.0
        captured = np.where(wet, np.minimum(np.maximum(raw_rain - 105.0, 0.0) * storage_capture, 42.0), 0.0)
        carried_storage += captured

        et0 = BASE_MONTHLY_ET0_MM[month] * (1.0 + 0.035 * temp_anom)
        potential_monthly_et = et0 * kc
        threshold = soil_capacity * 0.47
        irrigation_need = np.maximum(0.0, threshold + potential_monthly_et - (soil_water + effective_rain))
        irrigation_system = np.minimum(irrigation_budget, irrigation_need)
        irrigation_storage = np.minimum(carried_storage, np.maximum(0.0, irrigation_need - irrigation_system))
        carried_storage -= irrigation_storage
        gross_irrigation = irrigation_system + irrigation_storage
        net_irrigation = gross_irrigation * irrigation_efficiency

        adjusted_pet = potential_monthly_et * (1.0 - evap_reduction * implementation)
        available = soil_water + effective_rain + net_irrigation
        water_coefficient = np.clip(available / (soil_capacity * 0.58), 0.05, 1.0)
        month_aet = np.minimum(available, adjusted_pet * water_coefficient)
        soil_after = available - month_aet
        deep_drainage = np.maximum(0.0, soil_after - soil_capacity)
        soil_water = np.clip(soil_after - deep_drainage, 0.0, soil_capacity)

        potential_et += adjusted_pet
        actual_et += month_aet
        irrigation_applied += gross_irrigation
        effective_precipitation += effective_rain
        runoff_total += runoff
        monthly_satisfaction[:, month] = month_aet / np.maximum(adjusted_pet, 1e-9)

    monthly_temp = BASE_MONTHLY_TEMPERATURE_C[None, :] + temp_anom[:, None]
    stage_weights = np.stack([np.roll(STAGE_YIELD_WEIGHTS, shift) for shift in planting_shift])
    tolerance_width = 7.2 + col("heat_tolerance_shift_c") * implementation
    temperature_response = np.exp(-((monthly_temp - 27.0) / tolerance_width[:, None]) ** 2)
    temperature_factor = np.sum(stage_weights * temperature_response, axis=1)
    stage_water = np.sum(stage_weights * monthly_satisfaction, axis=1)
    water_factor = np.clip(stage_water ** (1.10 / col("water_use_efficiency_factor")), 0.12, 1.06)

    extreme_threshold = 35.4 + col("heat_tolerance_shift_c") * implementation
    heat_excess = np.maximum(monthly_temp - extreme_threshold[:, None], 0.0)
    heat_penalty = np.exp(-0.078 * np.sum(stage_weights * heat_excess ** 1.45, axis=1))
    flood_penalty = np.clip(1.0 - col("flood_index") * (1.0 - drainage * implementation) * 0.30, 0.78, 1.02)
    pest_penalty = np.clip(1.0 - col("pest_pressure") * (0.24 - 0.14 * col("pest_protection") * implementation), 0.86, 1.0)
    co2_factor = np.clip(1.0 + 0.000075 * (col("co2_ppm") - 420.0), 0.99, 1.045)
    management_factor = 1.0 + (col("potential_yield_factor") - 1.0) * implementation

    yield_t_ha = config.baseline_potential_yield_t_ha * management_factor * temperature_factor * water_factor * heat_penalty * flood_penalty * pest_penalty * co2_factor * col("model_structure_factor")
    yield_t_ha = np.clip(yield_t_ha, 26.0, 132.0)
    maturation_temp = monthly_temp[:, 8:12].mean(axis=1)
    sugar_content = np.clip(0.137 + 0.0028 * (26.0 - maturation_temp) - 0.010 * (1.0 - water_factor) - 0.006 * col("flood_index") + 0.004 * col("pest_protection") * implementation, 0.102, 0.158)
    sugar_yield = yield_t_ha * sugar_content

    consumptive_water = actual_et * 10.0
    blue_water = irrigation_applied * 10.0
    water_productivity = yield_t_ha * 1000.0 / np.maximum(consumptive_water, 1.0)
    energy_export = yield_t_ha * 0.0115 + col("energy_bonus_mwh_ha") * implementation
    cane_price = 43.5 * col("sugar_price_index")
    energy_price = 67.0 * col("energy_price_index")
    gross_revenue = yield_t_ha * cane_price + energy_export * energy_price
    variable_cost = 1120.0 * col("input_cost_index") + 6.8 * yield_t_ha + col("annual_opex_usd_ha") * implementation + blue_water * col("water_cost_usd_m3")
    annual_margin = gross_revenue - variable_cost
    discount = col("discount_rate")
    annuity = (1.0 - (1.0 + discount) ** (-config.planning_horizon_years)) / discount
    npv = -col("capex_usd_ha") * implementation + annual_margin * annuity

    pumping_emissions = blue_water * 0.00022 * (1.15 - 0.35 * irrigation_efficiency)
    residue_input = 3.95 + 0.012 * yield_t_ha + 0.55 * (col("input_cost_index") - 1.0)
    avoided = col("ghg_reduction_tco2e_ha") * implementation + 0.31 * energy_export
    ghg_balance = residue_input + pumping_emissions - avoided
    labor_days = 30.0 + 0.090 * yield_t_ha + col("labor_change_days_ha") * implementation
    complexity_burden = complexity * (1.0 - institutional)
    equity = np.clip(0.45 + 0.18 * np.tanh(annual_margin / 2500.0) + col("equity_bonus") * implementation + 0.16 * institutional - 0.22 * complexity_burden, 0.05, 0.98)
    irrigation_pressure = np.clip(blue_water / 3600.0, 0, 1.5)
    runoff_pressure = np.clip(runoff_total / 420.0, 0, 1.5)
    emission_pressure = np.clip((ghg_balance + 0.5) / 6.5, 0, 1.5)
    ecological_pressure = np.clip(0.42 * irrigation_pressure + 0.27 * runoff_pressure + 0.31 * emission_pressure - ecological_bonus * implementation, 0.02, 1.20)

    result = expanded[["scenario_id", "strategy_id", "name", "short_name", "pathway", "horizon"]].copy()
    result = result.rename(columns={"name": "strategy", "short_name": "strategy_short"})
    for name, values in {
        "implementation_factor": implementation,
        "temperature_factor": temperature_factor,
        "water_factor": water_factor,
        "heat_penalty": heat_penalty,
        "flood_penalty": flood_penalty,
        "pest_penalty": pest_penalty,
        "yield_t_ha": yield_t_ha,
        "sugar_content_fraction": sugar_content,
        "sugar_yield_t_ha": sugar_yield,
        "potential_et_mm": potential_et,
        "actual_et_mm": actual_et,
        "effective_precipitation_mm": effective_precipitation,
        "irrigation_mm": irrigation_applied,
        "blue_water_m3_ha": blue_water,
        "water_productivity_kg_m3": water_productivity,
        "annual_net_margin_usd_ha": annual_margin,
        "npv_usd_ha": npv,
        "energy_export_mwh_ha": energy_export,
        "ghg_balance_tco2e_ha": ghg_balance,
        "labor_days_ha": labor_days,
        "equity_score": equity,
        "ecological_pressure_index": ecological_pressure,
    }.items():
        result[name] = values
    # Retain scenario features for machine learning and decision analysis.
    scenario_columns = [c for c in expanded.columns if c in {
        "temperature_anomaly_c", "precipitation_change_fraction", "heatwave_days", "co2_ppm",
        "rainfall_cv", "sugar_price_index", "energy_price_index", "input_cost_index",
        "water_cost_usd_m3", "institutional_capacity", "adoption_fraction", "discount_rate",
        "pest_pressure", "flood_index", "climate_severity_index"
    }]
    for c in scenario_columns:
        result[c] = expanded[c].to_numpy()
    result["yield_failure"] = result["yield_t_ha"] < 70.0
    result["financial_failure"] = result["npv_usd_ha"] < 7000.0
    result["water_failure"] = result["water_productivity_kg_m3"] < 6.0
    result["joint_failure"] = result[["yield_failure", "financial_failure", "water_failure"]].any(axis=1)
    return result


def simulate_all_scenarios(scenarios: pd.DataFrame, config: ResearchConfig) -> pd.DataFrame:
    return simulate_expanded(expand_scenarios_and_strategies(scenarios), config)


# -----------------------------------------------------------------------------
# Section 5.1 outputs
# -----------------------------------------------------------------------------

def create_section_51_outputs(scenarios: pd.DataFrame, performance: pd.DataFrame, output_dir: Path, config: ResearchConfig) -> Dict[str, pd.DataFrame]:
    climate = scenarios.groupby(["pathway", "horizon"]).agg(
        n_scenarios=("scenario_id", "size"),
        mean_temperature_anomaly_c=("temperature_anomaly_c", "mean"),
        p05_temperature_anomaly_c=("temperature_anomaly_c", lambda x: np.quantile(x, .05)),
        p95_temperature_anomaly_c=("temperature_anomaly_c", lambda x: np.quantile(x, .95)),
        mean_precipitation_change=("precipitation_change_fraction", "mean"),
        p05_precipitation_change=("precipitation_change_fraction", lambda x: np.quantile(x, .05)),
        p95_precipitation_change=("precipitation_change_fraction", lambda x: np.quantile(x, .95)),
        mean_heatwave_days=("heatwave_days", "mean"),
        mean_severity_index=("climate_severity_index", "mean"),
    ).reset_index()
    climate.to_csv(output_dir / "Table_1_Climate_ensemble_summary.csv", index=False)

    summary = performance.groupby(["strategy_id", "strategy", "strategy_short"]).agg(
        mean_yield_t_ha=("yield_t_ha", "mean"),
        sd_yield_t_ha=("yield_t_ha", "std"),
        p05_yield_t_ha=("yield_t_ha", lambda x: np.quantile(x, .05)),
        p95_yield_t_ha=("yield_t_ha", lambda x: np.quantile(x, .95)),
        mean_water_productivity_kg_m3=("water_productivity_kg_m3", "mean"),
        mean_irrigation_mm=("irrigation_mm", "mean"),
        mean_npv_usd_ha=("npv_usd_ha", "mean"),
        p05_npv_usd_ha=("npv_usd_ha", lambda x: np.quantile(x, .05)),
        mean_ghg_balance_tco2e_ha=("ghg_balance_tco2e_ha", "mean"),
        mean_equity_score=("equity_score", "mean"),
        joint_failure_probability=("joint_failure", "mean"),
    ).reset_index().sort_values("mean_yield_t_ha", ascending=False)
    summary.to_csv(output_dir / "Table_2_Strategy_performance_summary.csv", index=False)

    fig, ax = plt.subplots(figsize=(8.4, 5.8))
    for (pathway, horizon), group in scenarios.groupby(["pathway", "horizon"]):
        ax.scatter(group["temperature_anomaly_c"], group["precipitation_change_fraction"] * 100, s=8, alpha=.45, label=f"{pathway}, {horizon}")
    ax.set_xlabel("Temperature anomaly (°C)")
    ax.set_ylabel("Annual precipitation change (%)")
    ax.set_title("Synthetic climate stress-test envelope")
    ax.legend(frameon=False, fontsize=7, ncol=3)
    save_figure(fig, output_dir / "Figure_1_Climate_scenario_envelope.png", config.figure_dpi)

    ordered = summary["strategy_short"].tolist()[::-1]
    data = [performance.loc[performance["strategy_short"] == s, "yield_t_ha"].to_numpy() for s in ordered]
    fig, ax = plt.subplots(figsize=(8.5, 5.8))
    ax.boxplot(data, vert=False, tick_labels=ordered, showfliers=False)
    ax.axvline(70, linestyle="--", linewidth=1, label="Yield adequacy threshold")
    ax.set_xlabel("Cane yield (t ha⁻¹)")
    ax.set_ylabel("Adaptation strategy")
    ax.set_title("Yield distributions across all plausible futures")
    ax.legend(frameon=False, fontsize=8)
    save_figure(fig, output_dir / "Figure_2_Yield_distributions_by_strategy.png", config.figure_dpi)

    # Response surfaces for current practice and the integrated package. Both the
    # individual panels and the publication-ready combined figure are generated
    # directly from the same saved scenario-strategy table.
    surface_data: Dict[str, pd.DataFrame] = {}
    surface_specs = [
        ("Current", "Figure_3a_Current_yield_response_surface.png"),
        ("Integrated package", "Figure_3b_Integrated_package_yield_response_surface.png"),
    ]
    for strategy_short, filename in surface_specs:
        subset = performance[performance["strategy_short"] == strategy_short].copy()
        t_bins = pd.cut(subset["temperature_anomaly_c"], bins=7)
        p_bins = pd.cut(subset["precipitation_change_fraction"] * 100, bins=7)
        pivot = subset.assign(t_bin=t_bins, p_bin=p_bins).pivot_table(
            index="p_bin", columns="t_bin", values="yield_t_ha", aggfunc="mean", observed=True
        )
        surface_data[strategy_short] = pivot
        fig, ax = plt.subplots(figsize=(7.2, 5.5))
        image = ax.imshow(pivot.values, aspect="auto", origin="lower")
        ax.set_xticks(range(len(pivot.columns)))
        ax.set_xticklabels([f"{i.mid:.1f}" for i in pivot.columns], rotation=45, ha="right", fontsize=8)
        ax.set_yticks(range(len(pivot.index)))
        ax.set_yticklabels([f"{i.mid:.0f}" for i in pivot.index], fontsize=8)
        ax.set_xlabel("Temperature anomaly midpoint (°C)")
        ax.set_ylabel("Precipitation change midpoint (%)")
        ax.set_title(f"Mean yield response surface: {strategy_short}")
        cbar = fig.colorbar(image, ax=ax)
        cbar.set_label("Mean cane yield (t ha⁻¹)")
        save_figure(fig, output_dir / filename, config.figure_dpi)

    fig, axes = plt.subplots(1, 2, figsize=(10.4, 4.45), sharey=True)
    for panel, (ax, (strategy_short, _)) in enumerate(zip(axes, surface_specs)):
        pivot = surface_data[strategy_short]
        image = ax.imshow(pivot.values, aspect="auto", origin="lower")
        ax.set_xticks(range(len(pivot.columns)))
        ax.set_xticklabels([f"{i.mid:.1f}" for i in pivot.columns], rotation=35, ha="right", fontsize=8)
        ax.set_yticks(range(len(pivot.index)))
        ax.set_yticklabels([f"{i.mid:.0f}" for i in pivot.index], fontsize=8)
        ax.set_xlabel("Temperature anomaly midpoint (°C)", fontsize=9)
        if panel == 0:
            ax.set_ylabel("Precipitation change midpoint (%)", fontsize=9)
        ax.set_title(f"{strategy_short}", fontsize=10)
        ax.text(-0.12, 1.04, f"({chr(97 + panel)})", transform=ax.transAxes, fontsize=9, fontweight="bold")
        cbar = fig.colorbar(image, ax=ax, fraction=0.046, pad=0.035)
        cbar.ax.tick_params(labelsize=7)
        cbar.set_label("Mean cane yield (t ha⁻¹)", fontsize=8)
    save_figure(fig, output_dir / "Figure_3_Combined_yield_response_surfaces.png", config.figure_dpi)

    fig, ax = plt.subplots(figsize=(8.2, 5.6))
    ax.scatter(summary["mean_irrigation_mm"], summary["mean_water_productivity_kg_m3"], s=70)
    for _, row in summary.iterrows():
        ax.annotate(row["strategy_short"], (row["mean_irrigation_mm"], row["mean_water_productivity_kg_m3"]), xytext=(4, 3), textcoords="offset points", fontsize=8)
    ax.set_xlabel("Mean irrigation application (mm year⁻¹)")
    ax.set_ylabel("Mean water productivity (kg m⁻³)")
    ax.set_title("Water-use trade-offs across adaptation strategies")
    save_figure(fig, output_dir / "Figure_4_Water_use_tradeoffs.png", config.figure_dpi)
    return {"climate_ensemble_summary": climate, "strategy_performance_summary": summary}


# -----------------------------------------------------------------------------
# Section 5.2 outputs: machine learning, conformal intervals, SHAP, sensitivity
# -----------------------------------------------------------------------------

SCENARIO_FEATURES = [
    "temperature_anomaly_c", "precipitation_change_fraction", "heatwave_days", "co2_ppm",
    "rainfall_cv", "sugar_price_index", "energy_price_index", "input_cost_index",
    "water_cost_usd_m3", "institutional_capacity", "adoption_fraction", "discount_rate",
    "pest_pressure", "flood_index", "climate_severity_index",
]
CATEGORICAL_FEATURES = ["strategy_short", "pathway", "horizon"]


def grouped_split(performance: pd.DataFrame, config: ResearchConfig) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    ids = performance["scenario_id"].drop_duplicates().to_numpy()
    rng = np.random.default_rng(config.random_seed + 101)
    rng.shuffle(ids)
    n_train = int(.60 * len(ids)); n_cal = int(.20 * len(ids))
    train_ids, cal_ids, test_ids = ids[:n_train], ids[n_train:n_train+n_cal], ids[n_train+n_cal:]
    return (
        performance[performance.scenario_id.isin(train_ids)].copy(),
        performance[performance.scenario_id.isin(cal_ids)].copy(),
        performance[performance.scenario_id.isin(test_ids)].copy(),
    )


def make_preprocessor() -> ColumnTransformer:
    return ColumnTransformer([
        ("numeric", StandardScaler(), SCENARIO_FEATURES),
        ("categorical", OneHotEncoder(handle_unknown="ignore", sparse_output=False), CATEGORICAL_FEATURES),
    ], remainder="drop", verbose_feature_names_out=False)


def regression_metrics(y: np.ndarray, pred: np.ndarray) -> Dict[str, float]:
    residual = y - pred
    mse = mean_squared_error(y, pred)
    denominator = np.sum((y - np.mean(y)) ** 2)
    return {
        "R2": r2_score(y, pred),
        "RMSE": math.sqrt(mse),
        "MAE": mean_absolute_error(y, pred),
        "MAPE_percent": float(np.mean(np.abs(residual) / np.maximum(np.abs(y), 1e-9)) * 100),
        "NSE": float(1 - np.sum(residual ** 2) / denominator),
        "Bias": float(np.mean(pred - y)),
    }


def create_section_52_outputs(performance: pd.DataFrame, output_dir: Path, config: ResearchConfig) -> Dict[str, pd.DataFrame]:
    train, calibration, test = grouped_split(performance, config)
    features = SCENARIO_FEATURES + CATEGORICAL_FEATURES
    X_train, y_train = train[features], train["yield_t_ha"].to_numpy()
    X_cal, y_cal = calibration[features], calibration["yield_t_ha"].to_numpy()
    X_test, y_test = test[features], test["yield_t_ha"].to_numpy()
    candidates = {
        "XGBoost": XGBRegressor(n_estimators=360, max_depth=5, learning_rate=.045, subsample=.82, colsample_bytree=.82, reg_lambda=1.2, objective="reg:squarederror", random_state=config.random_seed, n_jobs=max(1, os.cpu_count() or 1)),
        "Extra Trees": ExtraTreesRegressor(n_estimators=360, min_samples_leaf=2, max_features=.85, random_state=config.random_seed, n_jobs=-1),
        "Random Forest": RandomForestRegressor(n_estimators=320, min_samples_leaf=2, max_features=.82, random_state=config.random_seed, n_jobs=-1),
    }
    rows: List[Dict[str, Any]] = []
    fitted: Dict[str, Pipeline] = {}
    predictions: Dict[str, np.ndarray] = {}
    for name, estimator in candidates.items():
        pipe = Pipeline([("preprocessor", make_preprocessor()), ("model", estimator)])
        pipe.fit(X_train, y_train)
        pred = pipe.predict(X_test)
        rows.append({"model": name, **regression_metrics(y_test, pred)})
        fitted[name] = pipe; predictions[name] = pred
    model_table = pd.DataFrame(rows).sort_values("RMSE").reset_index(drop=True)
    model_table.to_csv(output_dir / "Table_3_Surrogate_model_performance.csv", index=False)
    best_name = str(model_table.iloc[0]["model"])
    best = fitted[best_name]
    test_pred = predictions[best_name]

    cal_pred = best.predict(X_cal)
    scores = np.abs(y_cal - cal_pred)
    alpha = 1.0 - config.target_conformal_coverage
    rank = math.ceil((len(scores) + 1) * (1 - alpha)) / len(scores)
    radius = float(np.quantile(scores, min(rank, 1.0), method="higher"))
    lower, upper = test_pred - radius, test_pred + radius
    covered = (y_test >= lower) & (y_test <= upper)
    conformal = pd.DataFrame([{
        "target_coverage": config.target_conformal_coverage,
        "empirical_test_coverage": float(covered.mean()),
        "mean_interval_width_t_ha": float(np.mean(upper - lower)),
        "conformal_radius_t_ha": radius,
        "calibration_sample_size": len(y_cal),
        "test_sample_size": len(y_test),
    }])
    conformal.to_csv(output_dir / "Table_4_Conformal_interval_diagnostics.csv", index=False)
    prediction_table = test[features].reset_index(drop=True).copy()
    prediction_table["observed_yield_t_ha"] = y_test
    prediction_table["predicted_yield_t_ha"] = test_pred
    prediction_table["residual_t_ha"] = y_test - test_pred
    prediction_table["lower_90_t_ha"] = lower
    prediction_table["upper_90_t_ha"] = upper
    prediction_table["covered_90"] = covered.astype(int)
    prediction_table.to_csv(output_dir / "Test_set_predictions_and_intervals.csv", index=False)

    fig, ax = plt.subplots(figsize=(6.3, 6.0))
    ax.scatter(y_test, test_pred, s=12, alpha=.36)
    lo, hi = min(y_test.min(), test_pred.min()), max(y_test.max(), test_pred.max())
    ax.plot([lo, hi], [lo, hi], linestyle="--", linewidth=1)
    ax.set_xlabel("Process-model yield (t ha⁻¹)")
    ax.set_ylabel("Surrogate-predicted yield (t ha⁻¹)")
    ax.set_title(f"Out-of-scenario validation of the {best_name} surrogate")
    save_figure(fig, output_dir / "Figure_5_Surrogate_parity_plot.png", config.figure_dpi)

    preprocessor = best.named_steps["preprocessor"]
    model = best.named_steps["model"]
    transformed = preprocessor.transform(X_train)
    feature_names = list(preprocessor.get_feature_names_out())
    sample_size = min(900, len(transformed))
    rng = np.random.default_rng(config.random_seed + 202)
    sample_idx = rng.choice(len(transformed), sample_size, replace=False)
    transformed_sample = transformed[sample_idx]
    if best_name == "XGBoost":
        explainer = shap.TreeExplainer(model)
        shap_values = np.asarray(explainer.shap_values(transformed_sample))
    else:
        explainer = shap.TreeExplainer(model)
        shap_values = np.asarray(explainer.shap_values(transformed_sample))
    importance = pd.DataFrame({"feature": feature_names, "mean_absolute_shap": np.mean(np.abs(shap_values), axis=0)}).sort_values("mean_absolute_shap", ascending=False)
    importance.to_csv(output_dir / "Table_5_SHAP_feature_importance.csv", index=False)
    top = importance.head(14).sort_values("mean_absolute_shap")
    fig, ax = plt.subplots(figsize=(8.2, 6.0))
    ax.barh(top["feature"], top["mean_absolute_shap"])
    ax.set_xlabel("Mean absolute SHAP value (t ha⁻¹)")
    ax.set_ylabel("Feature")
    ax.set_title("Global explanation of surrogate yield predictions")
    save_figure(fig, output_dir / "Figure_6_SHAP_global_importance.png", config.figure_dpi)

    continuous = [f for f in importance.feature if f in SCENARIO_FEATURES]
    dominant = continuous[0] if continuous else "climate_severity_index"
    j = feature_names.index(dominant)
    fig, ax = plt.subplots(figsize=(7.4, 5.2))
    ax.scatter(transformed_sample[:, j], shap_values[:, j], s=13, alpha=.45)
    ax.axhline(0, linewidth=.8)
    ax.set_xlabel(f"Standardized {dominant.replace('_', ' ')}")
    ax.set_ylabel("SHAP contribution to predicted yield (t ha⁻¹)")
    ax.set_title("Nonlinear marginal contribution of the dominant predictor")
    save_figure(fig, output_dir / "Figure_7_SHAP_dominant_feature_dependence.png", config.figure_dpi)

    interval_sample = prediction_table.sample(min(160, len(prediction_table)), random_state=config.random_seed).sort_values("predicted_yield_t_ha").reset_index(drop=True)
    x = np.arange(len(interval_sample))
    fig, ax = plt.subplots(figsize=(9.0, 5.2))
    ax.vlines(x, interval_sample.lower_90_t_ha, interval_sample.upper_90_t_ha, linewidth=.7, alpha=.65)
    ax.scatter(x, interval_sample.observed_yield_t_ha, s=12, label="Observed")
    ax.plot(x, interval_sample.predicted_yield_t_ha, linewidth=1, label="Predicted")
    ax.set_xlabel("Ordered test cases"); ax.set_ylabel("Cane yield (t ha⁻¹)")
    ax.set_title("Distribution-free 90% prediction intervals"); ax.legend(frameon=False)
    save_figure(fig, output_dir / "Figure_8_Conformal_prediction_intervals.png", config.figure_dpi)

    sensitivity = global_sensitivity_analysis(config)
    sensitivity.to_csv(output_dir / "Table_6_Global_sensitivity_indices.csv", index=False)
    sensitivity_panels: Dict[str, pd.DataFrame] = {}
    for target in ["yield_t_ha", "npv_usd_ha"]:
        subset = sensitivity[sensitivity.target == target].sort_values("total_order_index")
        sensitivity_panels[target] = subset
        positions = np.arange(len(subset)); width = .36
        fig, ax = plt.subplots(figsize=(8.2, 5.8))
        ax.barh(positions - width/2, subset.first_order_index, height=width, label="First-order")
        ax.barh(positions + width/2, subset.total_order_index, height=width, label="Total-order")
        ax.set_yticks(positions); ax.set_yticklabels(subset.parameter.str.replace("_", " "))
        ax.set_xlabel("Jansen sensitivity index"); ax.set_ylabel("Uncertain input")
        label = "yield" if target == "yield_t_ha" else "net present value"
        ax.set_title(f"Global sensitivity of integrated-package {label}"); ax.legend(frameon=False)
        save_figure(fig, output_dir / f"Figure_9_{target}_sensitivity.png", config.figure_dpi)

    fig, axes = plt.subplots(1, 2, figsize=(10.5, 4.55))
    for panel, (ax, target) in enumerate(zip(axes, ["yield_t_ha", "npv_usd_ha"])):
        subset = sensitivity_panels[target]
        positions = np.arange(len(subset)); width = .34
        ax.barh(positions - width/2, subset.first_order_index, height=width, label="First-order")
        ax.barh(positions + width/2, subset.total_order_index, height=width, label="Total-order")
        ax.set_yticks(positions)
        ax.set_yticklabels(subset.parameter.str.replace("_", " "), fontsize=7.5)
        ax.set_xlabel("Jansen sensitivity index", fontsize=9)
        if panel == 0:
            ax.set_ylabel("Uncertain input", fontsize=9)
        label = "Yield" if target == "yield_t_ha" else "Net present value"
        ax.set_title(label, fontsize=10)
        ax.text(-0.15, 1.04, f"({chr(97 + panel)})", transform=ax.transAxes, fontsize=9, fontweight="bold")
        ax.tick_params(axis="x", labelsize=8)
    axes[1].legend(frameon=False, fontsize=8, loc="lower right")
    save_figure(fig, output_dir / "Figure_9_Combined_global_sensitivity.png", config.figure_dpi)
    return {"surrogate_model_performance": model_table, "conformal_diagnostics": conformal, "shap_feature_importance": importance, "global_sensitivity_indices": sensitivity}


SENSITIVITY_BOUNDS: Mapping[str, Tuple[float, float]] = {
    "temperature_anomaly_c": (.6, 5.3),
    "precipitation_change_fraction": (-.34, .12),
    "heatwave_days": (4, 66),
    "co2_ppm": (420, 790),
    "rainfall_cv": (.14, .47),
    "sugar_price_index": (.78, 1.36),
    "input_cost_index": (.86, 1.36),
    "water_cost_usd_m3": (.03, .12),
    "institutional_capacity": (.30, .95),
    "pest_pressure": (.02, .33),
}


def sensitivity_scenarios(matrix: np.ndarray, names: Sequence[str], offset: int) -> pd.DataFrame:
    d = pd.DataFrame(matrix, columns=names)
    institution = d.institutional_capacity.to_numpy()
    precip = d.precipitation_change_fraction.to_numpy()
    rain_cv = d.rainfall_cv.to_numpy()
    d.insert(0, "scenario_id", np.arange(offset, offset + len(d)))
    d["pathway"] = "Sensitivity envelope"; d["horizon"] = 2075
    d["energy_price_index"] = 1.0
    d["adoption_fraction"] = np.clip(.26 + .70 * institution, .28, .96)
    d["discount_rate"] = np.clip(.104 - .050 * institution, .04, .10)
    d["flood_index"] = np.clip(.35 * rain_cv + .50 * np.maximum(precip, 0) + .04, 0, .48)
    d["model_structure_factor"] = 1.0
    d["climate_severity_index"] = normalized_severity(d.temperature_anomaly_c.to_numpy(), precip, d.heatwave_days.to_numpy(), rain_cv)
    return d


def evaluate_sensitivity(matrix: np.ndarray, names: Sequence[str], config: ResearchConfig, offset: int) -> Tuple[np.ndarray, np.ndarray]:
    scenarios = sensitivity_scenarios(matrix, names, offset)
    strategy = strategy_frame().query("strategy_id == 6")
    expanded = scenarios.assign(_join_key=1).merge(strategy.assign(_join_key=1), on="_join_key").drop(columns="_join_key")
    result = simulate_expanded(expanded, config)
    return result.yield_t_ha.to_numpy(), result.npv_usd_ha.to_numpy()


def global_sensitivity_analysis(config: ResearchConfig) -> pd.DataFrame:
    names = list(SENSITIVITY_BOUNDS); d = len(names)
    power = math.ceil(math.log2(config.sobol_base_size))
    unit = qmc.Sobol(d=2*d, scramble=True, seed=config.random_seed + 303).random_base2(power)[:config.sobol_base_size]
    lower = np.array([SENSITIVITY_BOUNDS[n][0] for n in names]); upper = np.array([SENSITIVITY_BOUNDS[n][1] for n in names])
    A = qmc.scale(unit[:, :d], lower, upper); B = qmc.scale(unit[:, d:], lower, upper)
    yA_y, yA_n = evaluate_sensitivity(A, names, config, 700000)
    yB_y, yB_n = evaluate_sensitivity(B, names, config, 800000)
    mixed: List[Tuple[np.ndarray, np.ndarray]] = []
    for i in range(d):
        AB = A.copy(); AB[:, i] = B[:, i]
        mixed.append(evaluate_sensitivity(AB, names, config, 900000 + i*len(A)))
    records: List[Dict[str, Any]] = []
    for target, yA, yB, pos in [("yield_t_ha", yA_y, yB_y, 0), ("npv_usd_ha", yA_n, yB_n, 1)]:
        variance = np.var(np.concatenate([yA, yB]), ddof=1)
        for i, name in enumerate(names):
            yAB = mixed[i][pos]
            first = 1 - np.mean((yB - yAB)**2)/(2*variance)
            total = np.mean((yA - yAB)**2)/(2*variance)
            records.append({"target": target, "parameter": name, "first_order_index": float(np.clip(first, -.1, 1.1)), "total_order_index": float(np.clip(total, 0, 1.2))})
    return pd.DataFrame(records)


# -----------------------------------------------------------------------------
# Section 5.3 outputs: MCDA, regret, Pareto efficiency, pathways
# -----------------------------------------------------------------------------

DECISION_CRITERIA: Tuple[Tuple[str, bool], ...] = (
    ("yield_t_ha", True), ("npv_usd_ha", True), ("water_productivity_kg_m3", True),
    ("energy_export_mwh_ha", True), ("equity_score", True),
    ("ghg_balance_tco2e_ha", False), ("ecological_pressure_index", False),
)
BASE_WEIGHTS = np.array([.18, .20, .14, .10, .13, .12, .13])


def scenario_normalized_criteria(performance: pd.DataFrame) -> pd.DataFrame:
    parts = []
    for _, group in performance.groupby("scenario_id", sort=False):
        out = group[["scenario_id", "strategy_id", "strategy", "strategy_short"]].copy()
        for criterion, benefit in DECISION_CRITERIA:
            out[criterion] = minmax(group[criterion].reset_index(drop=True), benefit).to_numpy()
        parts.append(out)
    return pd.concat(parts, ignore_index=True)


def robust_multicriteria_analysis(performance: pd.DataFrame, config: ResearchConfig) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    normalized = scenario_normalized_criteria(performance)
    criteria = [c for c, _ in DECISION_CRITERIA]
    normalized["base_weight_utility"] = normalized[criteria].to_numpy() @ BASE_WEIGHTS
    normalized["best_scenario_utility"] = normalized.groupby("scenario_id").base_weight_utility.transform("max")
    normalized["regret"] = normalized.best_scenario_utility - normalized.base_weight_utility
    perf = performance.merge(normalized[["scenario_id", "strategy_id", "base_weight_utility", "regret"]], on=["scenario_id", "strategy_id"])

    profiles = normalized.groupby(["strategy_id", "strategy", "strategy_short"])[criteria].mean().reset_index().sort_values("strategy_id").reset_index(drop=True)
    rng = np.random.default_rng(config.random_seed + 777)
    sampled_weights = rng.dirichlet(BASE_WEIGHTS * 54.0, size=config.n_weight_samples)
    utilities = profiles[criteria].to_numpy() @ sampled_weights.T
    ranks = stats.rankdata(-utilities, axis=0, method="average")
    accept_records = []
    for i, row in profiles.iterrows():
        accept_records.append({
            "strategy_id": int(row.strategy_id), "strategy": row.strategy, "strategy_short": row.strategy_short,
            "probability_rank_1": float(np.mean(ranks[i] == 1)), "probability_top_3": float(np.mean(ranks[i] <= 3)),
            "median_rank": float(np.median(ranks[i])), "mean_rank": float(np.mean(ranks[i])),
            "mean_weighted_utility": float(np.mean(utilities[i])), "p05_weighted_utility": float(np.quantile(utilities[i], .05)),
            "p95_weighted_utility": float(np.quantile(utilities[i], .95)),
        })
    accept = pd.DataFrame(accept_records)
    robust = perf.groupby(["strategy_id", "strategy", "strategy_short"]).agg(
        expected_utility=("base_weight_utility", "mean"), p05_utility=("base_weight_utility", lambda x: np.quantile(x, .05)),
        maximum_regret=("regret", "max"), p90_regret=("regret", lambda x: np.quantile(x, .90)),
        expected_yield_t_ha=("yield_t_ha", "mean"), yield_reliability=("yield_failure", lambda x: 1-x.mean()),
        financial_reliability=("financial_failure", lambda x: 1-x.mean()), joint_reliability=("joint_failure", lambda x: 1-x.mean()),
        expected_npv_usd_ha=("npv_usd_ha", "mean"), p05_npv_usd_ha=("npv_usd_ha", lambda x: np.quantile(x, .05)),
        expected_water_productivity=("water_productivity_kg_m3", "mean"), expected_ghg_balance=("ghg_balance_tco2e_ha", "mean"),
        expected_equity=("equity_score", "mean"), expected_ecological_pressure=("ecological_pressure_index", "mean"),
    ).reset_index().merge(accept, on=["strategy_id", "strategy", "strategy_short"])
    robust["robustness_score"] = .24*minmax(robust.expected_utility, True) + .18*minmax(robust.p05_utility, True) + .20*minmax(robust.joint_reliability, True) + .16*minmax(robust.p90_regret, False) + .12*minmax(robust.probability_top_3, True) + .10*minmax(robust.p05_npv_usd_ha, True)
    robust["robust_rank"] = robust.robustness_score.rank(ascending=False, method="min").astype(int)
    robust = robust.sort_values(["robust_rank", "strategy_id"]).reset_index(drop=True)
    objectives = robust[["expected_npv_usd_ha", "joint_reliability", "expected_ghg_balance", "expected_ecological_pressure"]].to_numpy()
    transformed = objectives.copy(); transformed[:, 2:] *= -1
    efficient = np.ones(len(transformed), dtype=bool)
    for i in range(len(transformed)):
        for j in range(len(transformed)):
            if i != j and np.all(transformed[j] >= transformed[i]) and np.any(transformed[j] > transformed[i]):
                efficient[i] = False; break
    robust["pareto_efficient"] = efficient
    return robust, accept, perf


def derive_adaptive_pathways(perf: pd.DataFrame) -> pd.DataFrame:
    classes = pd.cut(perf.climate_severity_index, bins=[-.001, .30, .48, .66, 1.001], labels=["Low", "Moderate", "High", "Severe"])
    work = perf.assign(severity_class=classes)
    candidate = work.groupby(["severity_class", "strategy_id", "strategy", "strategy_short"], observed=True).agg(
        mean_utility=("base_weight_utility", "mean"), joint_reliability=("joint_failure", lambda x: 1-x.mean()),
        mean_yield_t_ha=("yield_t_ha", "mean"), mean_npv_usd_ha=("npv_usd_ha", "mean"), mean_irrigation_mm=("irrigation_mm", "mean"),
    ).reset_index()
    # Adaptive pathways impose staged feasibility rather than allowing the most capital-intensive
    # package to be selected immediately. Low-severity states admit low-regret measures; moderate
    # states admit integrated investment; high and severe states admit the full portfolio.
    eligible_by_class = {
        "Low": {0, 1, 3, 4, 7},
        "Moderate": {0, 1, 2, 3, 4, 6, 7},
        "High": set(range(9)),
        "Severe": set(range(9)),
    }
    selected_rows = []
    for severity in ["Low", "Moderate", "High", "Severe"]:
        subset = candidate[(candidate["severity_class"].astype(str) == severity) & candidate["strategy_id"].isin(eligible_by_class[severity])]
        selected_rows.append(subset.sort_values("mean_utility", ascending=False).iloc[0])
    chosen = pd.DataFrame(selected_rows).reset_index(drop=True)
    previous = ["Current practice"] + chosen.strategy.tolist()[:-1]
    triggers = [
        "Maintain monitoring; act if three-year yield reliability falls below 0.90.",
        "Escalate when temperature anomaly exceeds 1.8 °C or seasonal rainfall deficit exceeds 10%.",
        "Escalate when heatwave exposure exceeds 32 days or irrigation demand exceeds 220 mm year⁻¹.",
        "Activate a transformational contingency when joint reliability falls below 0.80 or tail NPV becomes unacceptable.",
    ]
    return pd.DataFrame({
        "severity_class": chosen.severity_class.astype(str), "recommended_action": chosen.strategy.to_numpy(), "previous_action": previous,
        "mean_utility": chosen.mean_utility.to_numpy(), "joint_reliability": chosen.joint_reliability.to_numpy(),
        "mean_yield_t_ha": chosen.mean_yield_t_ha.to_numpy(), "mean_npv_usd_ha": chosen.mean_npv_usd_ha.to_numpy(),
        "mean_irrigation_mm": chosen.mean_irrigation_mm.to_numpy(), "adaptation_trigger": triggers,
    })


def create_pathway_figure(pathways: pd.DataFrame, path: Path, dpi: int) -> None:
    """Create a compact, publication-readable adaptive-pathway diagram."""
    short_actions = {
        "Low": "Soil conservation",
        "Moderate": "Integrated package",
        "High": "Circular package",
        "Severe": "Circular package\n+ transformation",
    }
    short_triggers = {
        "Low": "Monitor three-year\nyield reliability",
        "Moderate": "ΔT > 1.8 °C or\nrainfall deficit > 10%",
        "High": "Heatwaves > 32 d or\nirrigation > 220 mm",
        "Severe": "Joint reliability < 0.80\nor tail NPV unacceptable",
    }
    fig, ax = plt.subplots(figsize=(8.5, 3.25))
    ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis("off")
    x_positions = np.linspace(0.12, 0.88, 4)
    box_width, box_height, y0 = 0.205, 0.54, 0.20
    for i, (_, row) in enumerate(pathways.iterrows()):
        severity = str(row.severity_class)
        x0 = x_positions[i] - box_width / 2
        box = FancyBboxPatch(
            (x0, y0), box_width, box_height,
            boxstyle="round,pad=0.012,rounding_size=0.02",
            linewidth=1.2, facecolor="white", edgecolor="black",
        )
        ax.add_patch(box)
        ax.text(x_positions[i], y0 + 0.43, f"{severity.upper()} STRESS", ha="center", va="center", fontsize=10.5, fontweight="bold")
        ax.text(x_positions[i], y0 + 0.30, short_actions[severity], ha="center", va="center", fontsize=9.0)
        ax.plot([x0 + 0.018, x0 + box_width - 0.018], [y0 + 0.21, y0 + 0.21], linewidth=0.8)
        ax.text(x_positions[i], y0 + 0.105, short_triggers[severity], ha="center", va="center", fontsize=8.1, linespacing=1.15)
        if i < len(x_positions) - 1:
            ax.annotate(
                "", xy=(x_positions[i + 1] - box_width / 2 - 0.012, y0 + box_height / 2),
                xytext=(x_positions[i] + box_width / 2 + 0.012, y0 + box_height / 2),
                arrowprops={"arrowstyle": "-|>", "linewidth": 1.4},
            )
    ax.set_title("Adaptive pathway from low-regret action to transformational contingency", fontsize=11.5, pad=8)
    save_figure(fig, path, dpi)


def create_section_53_outputs(performance: pd.DataFrame, output_dir: Path, config: ResearchConfig) -> Dict[str, pd.DataFrame]:
    robust, accept, perf = robust_multicriteria_analysis(performance, config)
    robust.to_csv(output_dir / "Table_7_Robust_strategy_ranking.csv", index=False)
    accept.to_csv(output_dir / "Rank_acceptability_probabilities.csv", index=False)
    pathways = derive_adaptive_pathways(perf)
    pathways.to_csv(output_dir / "Table_8_Adaptive_pathway_triggers.csv", index=False)

    fig, ax = plt.subplots(figsize=(8.2, 5.6))
    ax.scatter(robust.expected_utility, robust.joint_reliability, s=90)
    robustness_offsets = {
        "Circular package": (5, 4), "Integrated package": (5, 4), "Soil conservation": (5, 4),
        "Water storage": (5, 4), "Deficit irrigation": (5, 4), "Cultivar": (5, 9),
        "Harvest coordination": (5, -13), "Planting shift": (5, -13), "Current": (-23, 6),
    }
    for _, row in robust.iterrows():
        ax.annotate(
            row.strategy_short, (row.expected_utility, row.joint_reliability),
            xytext=robustness_offsets.get(row.strategy_short, (5, 4)), textcoords="offset points", fontsize=8,
        )
    ax.margins(x=0.08, y=0.10)
    ax.set_xlabel("Expected multicriteria utility"); ax.set_ylabel("Joint reliability"); ax.yaxis.set_major_formatter(PercentFormatter(1))
    ax.set_title("Robustness and expected performance")
    save_figure(fig, output_dir / "Figure_10_Robustness_expected_utility.png", config.figure_dpi)

    fig, ax = plt.subplots(figsize=(8.2, 5.6))
    ax.scatter(robust.expected_ghg_balance, robust.expected_npv_usd_ha, s=np.where(robust.pareto_efficient, 135, 60))
    pareto_label_specs = {
        "Circular package": ((5, 4), "left"), "Integrated package": ((5, 4), "left"),
        "Soil conservation": ((5, 4), "left"), "Water storage": ((5, 4), "left"),
        "Deficit irrigation": ((5, 4), "left"), "Cultivar": ((-8, 8), "right"),
        "Harvest coordination": ((-8, 7), "right"), "Planting shift": ((8, -17), "left"),
        "Current": ((8, 11), "left"),
    }
    for _, row in robust.iterrows():
        offset, horizontal_alignment = pareto_label_specs.get(row.strategy_short, ((5, 4), "left"))
        ax.annotate(
            row.strategy_short, (row.expected_ghg_balance, row.expected_npv_usd_ha),
            xytext=offset, textcoords="offset points", fontsize=8, ha=horizontal_alignment,
        )
    ax.margins(x=0.08, y=0.10)
    ax.set_xlabel("Expected net GHG balance (t CO₂e ha⁻¹ year⁻¹; lower is better)"); ax.set_ylabel("Expected NPV (USD ha⁻¹)")
    ax.set_title("Economic-mitigation trade-off and finite-strategy Pareto set")
    save_figure(fig, output_dir / "Figure_11_NPV_GHG_Pareto_tradeoff.png", config.figure_dpi)

    regret = perf.groupby(["strategy_short", "pathway", "horizon"]).regret.mean().reset_index()
    regret["future_family"] = regret.pathway + "\n" + regret.horizon.astype(str)
    pivot = regret.pivot(index="strategy_short", columns="future_family", values="regret")
    fig, ax = plt.subplots(figsize=(9.4, 6.2)); image = ax.imshow(pivot.values, aspect="auto")
    ax.set_xticks(range(len(pivot.columns))); ax.set_xticklabels(pivot.columns, rotation=45, ha="right", fontsize=8)
    ax.set_yticks(range(len(pivot.index))); ax.set_yticklabels(pivot.index, fontsize=8)
    ax.set_xlabel("Climate pathway and horizon"); ax.set_ylabel("Adaptation strategy"); ax.set_title("Mean opportunity-loss regret across future families")
    cbar = fig.colorbar(image, ax=ax); cbar.set_label("Regret in normalized utility units")
    save_figure(fig, output_dir / "Figure_12_Regret_heatmap.png", config.figure_dpi)

    rank_plot = robust.sort_values("probability_rank_1")
    fig, ax = plt.subplots(figsize=(8.2, 5.7)); ax.barh(rank_plot.strategy_short, rank_plot.probability_rank_1)
    ax.xaxis.set_major_formatter(PercentFormatter(1)); ax.set_xlabel("Probability of first rank under uncertain stakeholder weights")
    ax.set_ylabel("Adaptation strategy"); ax.set_title("Rank acceptability under preference uncertainty")
    save_figure(fig, output_dir / "Figure_13_Rank_acceptability.png", config.figure_dpi)
    create_pathway_figure(pathways, output_dir / "Figure_14_Adaptive_pathway_map.png", config.figure_dpi)
    return {"robust_strategy_ranking": robust, "rank_acceptability": accept, "adaptive_pathway_triggers": pathways}


# -----------------------------------------------------------------------------
# Workbook, README, and manifest
# -----------------------------------------------------------------------------

def write_excel_workbook(root: Path, tables: Mapping[str, pd.DataFrame]) -> Path:
    path = root / "Integrated_Climate_Futures_Sugarcane_Results.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, table in tables.items(): table.to_excel(writer, sheet_name=sheet[:31], index=False)
    workbook = load_workbook(path)
    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="4B2E83"); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in ws.columns:
            letter = column[0].column_letter
            width = min(34, max(10, max(len(str(c.value)) if c.value is not None else 0 for c in column) + 2))
            ws.column_dimensions[letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path); return path


def stable_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024*1024), b""): h.update(chunk)
    return h.hexdigest()


def write_readme(root: Path, config: ResearchConfig, runtime: float) -> Path:
    text = f"""# Integrated Climate Futures and Decision Support for Resilient Sugarcane

This reproducibility package contains the complete synthetic worked case used in the chapter.
It is intended to demonstrate auditable integration, not to prescribe management in a real region.
Local calibration, independent field validation, and stakeholder review are required before use.

## Output structure

- `5.1_Climate_and_System_Response`: scenario envelope, process-model response, and resource trade-offs.
- `5.2_Predictive_Performance_and_Uncertainty`: surrogate validation, SHAP explanations, conformal intervals, and Jansen sensitivity indices.
- `5.3_Robust_Decisions_and_Adaptive_Pathways`: multicriteria robustness, regret, Pareto trade-offs, preference uncertainty, and adaptive triggers.
- `data`: synthetic scenarios and scenario-strategy performance records.

## Reproduction

```bash
python integrated_climate_futures_sugarcane.py --output-dir "PATH_TO_RESULTS"
```

Use `--auto-install` only when package installation is permitted. The deterministic seed is
{config.random_seed}; the ensemble contains {config.n_scenarios} futures; the Sobol base size is
{config.sobol_base_size}; and the measured runtime for this run was {runtime:.2f} seconds.
"""
    path = root / "README.md"; path.write_text(text, encoding="utf-8"); return path


def write_manifest(root: Path, config: ResearchConfig, runtime: float) -> Path:
    files = sorted(p for p in root.rglob("*") if p.is_file() and p.name != "reproducibility_manifest.json")
    manifest = {
        "chapter": "Integrated Climate Futures and Decision Support for Resilient Sugarcane",
        "created_utc": pd.Timestamp.utcnow().isoformat(),
        "runtime_seconds": runtime,
        "configuration": asdict(config),
        "library_versions": {name: importlib.import_module(module).__version__ for module, name in [("numpy","numpy"),("pandas","pandas"),("scipy","scipy"),("sklearn","scikit-learn"),("matplotlib","matplotlib"),("xgboost","xgboost"),("shap","shap")]},
        "files": [{"relative_path": str(p.relative_to(root)), "size_bytes": p.stat().st_size, "sha256": stable_hash(p)} for p in files],
    }
    path = root / "reproducibility_manifest.json"; path.write_text(json.dumps(manifest, indent=2), encoding="utf-8"); return path


# -----------------------------------------------------------------------------
# Command-line workflow
# -----------------------------------------------------------------------------

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the integrated climate-futures and robust sugarcane decision-support case.")
    parser.add_argument("--output-dir", type=Path, default=default_output_dir(), help="Directory for all outputs.")
    parser.add_argument("--n-scenarios", type=int, default=1800, help="Balanced number of plausible futures; must be divisible by nine.")
    parser.add_argument("--seed", type=int, default=4901, help="Deterministic random seed.")
    parser.add_argument("--weight-samples", type=int, default=3000, help="Dirichlet samples for rank acceptability.")
    parser.add_argument("--sobol-base-size", type=int, default=512, help="Base size for Jansen sensitivity estimation.")
    parser.add_argument("--auto-install", action="store_true", help="Attempt installation of missing dependencies.")
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    if args.n_scenarios < 450 or args.n_scenarios % 9 != 0: raise ValueError("--n-scenarios must be at least 450 and divisible by nine.")
    if args.weight_samples < 500: raise ValueError("--weight-samples must be at least 500.")
    if args.sobol_base_size < 128: raise ValueError("--sobol-base-size must be at least 128.")
    config = ResearchConfig(random_seed=args.seed, n_scenarios=args.n_scenarios, n_weight_samples=args.weight_samples, sobol_base_size=args.sobol_base_size, output_dir=str(args.output_dir.resolve()))
    paths = ensure_output_structure(args.output_dir.resolve())
    started = time.perf_counter()
    print("[1/6] Generating the balanced climate and socioeconomic scenario ensemble...", flush=True)
    scenarios = generate_scenario_ensemble(config); scenarios.to_csv(paths["data"] / "synthetic_climate_scenarios.csv", index=False)
    print("[2/6] Running the monthly process-informed system model...", flush=True)
    performance = simulate_all_scenarios(scenarios, config); performance.to_csv(paths["data"] / "scenario_strategy_performance.csv", index=False)
    print("[3/6] Creating Section 5.1 evidence...", flush=True)
    t51 = create_section_51_outputs(scenarios, performance, paths["s51"], config)
    print("[4/6] Creating Section 5.2 predictive and uncertainty evidence...", flush=True)
    t52 = create_section_52_outputs(performance, paths["s52"], config)
    print("[5/6] Creating Section 5.3 robust-decision evidence...", flush=True)
    t53 = create_section_53_outputs(performance, paths["s53"], config)
    tables = {
        "Climate ensemble": t51["climate_ensemble_summary"], "Strategy summary": t51["strategy_performance_summary"],
        "ML performance": t52["surrogate_model_performance"], "Conformal diagnostics": t52["conformal_diagnostics"],
        "SHAP importance": t52["shap_feature_importance"], "Sensitivity indices": t52["global_sensitivity_indices"],
        "Robust ranking": t53["robust_strategy_ranking"], "Rank acceptability": t53["rank_acceptability"],
        "Adaptive pathways": t53["adaptive_pathway_triggers"],
    }
    workbook = write_excel_workbook(paths["root"], tables)
    runtime = time.perf_counter() - started
    readme = write_readme(paths["root"], config, runtime)
    manifest = write_manifest(paths["root"], config, runtime)
    print("[6/6] Reproducibility package completed.", flush=True)
    print(f"Output directory: {paths['root']}")
    print(f"Workbook: {workbook}\nREADME: {readme}\nManifest: {manifest}\nRuntime: {runtime:.2f} seconds")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
